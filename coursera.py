import requests
from bs4 import BeautifulSoup, NavigableString
import argparse
from openpyxl import Workbook


def fetch_xml():
    return requests.get("https://www.coursera.org/sitemap~www~courses.xml").text


def fetch_html(course_url):
    return requests.get(course_url).text


def get_courses_list(xml):
    xml_soup = BeautifulSoup(xml, features="lxml-xml")
    return [course.contents[0] for course in xml_soup.find_all("loc")]


def get_course_info(html):
    html_soup = BeautifulSoup(html, features="lxml")
    return get_course_name(html_soup), get_course_lang(html_soup), \
           get_course_week_quantity(html_soup), get_course_average_rating(html_soup)


def title_worksheet(worksheet):
    worksheet.cell(column=1, row=1, value="Название курса")
    worksheet.cell(column=2, row=1, value="Ссылка")
    worksheet.cell(column=3, row=1, value="Языки")
    worksheet.cell(column=4, row=1, value="Продолжительность")
    worksheet.cell(column=5, row=1, value="Рейтинг")


def output_courses_info_to_xlsx(filepath, courses_list):
    workbook = Workbook()
    worksheet = workbook.active
    title_worksheet(worksheet)
    list_cell_id = 0
    for row in range(2, len(courses_list)+2):  # first row is titles
        course_info = get_course_info(fetch_html(courses_list[list_cell_id]))
        course_name = course_info[0].encode('raw-unicode-escape').decode('utf-8-sig')
        course_link = courses_list[list_cell_id]
        course_lang = course_info[1]
        course_weeks = course_info[2]
        course_rating = course_info[3]
        worksheet.cell(column=1, row=row, value=course_name)
        worksheet.cell(column=2, row=row, value=course_link)
        if not type(course_info[1]) == tuple:  # course has no subtitles
            worksheet.cell(column=3, row=row, value=course_lang)
        else:                           # course has got subtitles
            worksheet.cell(column=3, row=row, value="{}, субтитры: {}".format(course_lang[0], course_lang[1]))
        if course_weeks:
            worksheet.cell(column=4, row=row, value="{} недель".format(course_weeks))
        else:
            worksheet.cell(column=4, row=row, value="Нет информации")
        if course_rating:
            if type(course_rating) == NavigableString:
                worksheet.cell(column=5, row=row, value=course_rating.replace("stars", "из 5"))
            else:  # unusual html markup on page
                rating = course_rating.contents[1][6:8]  # get only numbers for output in Russian
                worksheet.cell(column=5, row=row, value="{} из 5".format(rating))
        else:
            worksheet.cell(column=5, row=row, value="Оценок еще нет")
        list_cell_id += 1
    workbook.save(filename=filepath)


def get_course_name(html_soup):
    name = html_soup.find("h1")
    if name:
        return name.contents[0]


def get_course_lang(html_soup):
    lang = html_soup.find('div', attrs={'class': 'rc-Language'})
    lang_no_subtitles_len = 3
    if not lang:
        return None
    if len(lang) == lang_no_subtitles_len:
        return lang.contents[1]
    else:
        return lang.contents[1], lang.contents[3].contents[3]  # if course has got subtitles, tuple returns


def get_course_week_quantity(html_soup):
    weeks = html_soup.find('div', attrs={'class': 'rc-WeekView'})
    if weeks:
        return len(weeks)


def get_course_average_rating(html_soup):
    rating = html_soup.find('div', attrs={'class': 'ratings-text'})
    if rating:
        return rating.contents[0]


def parse_args():
    parser = argparse.ArgumentParser()
    parser.add_argument("output", type=str, help="xlsx-файл в который будут выгружены данные")
    parser.add_argument("-q", "--quantity", type=int, help="Колличество спарсеных курсов. Изначально – 20", default=20)
    return parser.parse_args()


if __name__ == '__main__':
    output_courses_info_to_xlsx(parse_args().output, get_courses_list(fetch_xml()))
