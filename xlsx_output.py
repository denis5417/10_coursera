from openpyxl import Workbook
import courses_data_prettify as cds
import courses_data_get as cdg
from fetch_webpage import fetch_webpage


def title_worksheet(worksheet):
    worksheet.cell(column=1, row=1, value="Название курса")
    worksheet.cell(column=2, row=1, value="Ссылка")
    worksheet.cell(column=3, row=1, value="Языки")
    worksheet.cell(column=4, row=1, value="Продолжительность")
    worksheet.cell(column=5, row=1, value="Рейтинг")


def output_courses_info_to_xlsx(filepath, courses_list):
    workbook = Workbook()
    worksheet = workbook.active
    title_worksheet(worksheet)
    courses_cell_id = 0
    for row in range(2, len(courses_list) + 2):
        course_info = cdg.agregate_course_info(fetch_webpage(courses_list[courses_cell_id]))
        fill_row(worksheet, row, course_info, courses_list[courses_cell_id])
        courses_cell_id += 1
    workbook.save(filename=filepath)


def fill_row(worksheet, row, course_info, link):
    worksheet.cell(column=1, row=row, value=cds.prettify_course_name(course_info['Name']))
    worksheet.cell(column=2, row=row, value=link)
    worksheet.cell(column=3, row=row, value=cds.prettify_course_languages(course_info['Languages']))
    worksheet.cell(column=4, row=row, value=cds.prettify_course_weeks_count(course_info['Weeks']))
    worksheet.cell(column=5, row=row, value=cds.prettify_course_rating(course_info['Rating']))
